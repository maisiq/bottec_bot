import os
import re

from openpyxl import Workbook, load_workbook

ORDERS_FILE = 'orders.xlsx'


def update_order_on_refund(provider_payment_charge_id):
    wb = load_workbook(ORDERS_FILE)
    ws = wb.active

    # В openpyxl индексирование начинается с 1
    # Payment Charge ID (2-й столбец) и Paid (7-й столбец)
    payment_charge_col = 2
    paid_col = 7

    for row in range(2, ws.max_row + 1):  # Пропускаем заголовок
        if ws.cell(row=row, column=payment_charge_col).value == provider_payment_charge_id:
            ws.cell(row=row, column=paid_col).value = 'Возврат средств'

    wb.save('orders-updated.xlsx')


def save_order(
    client_username,
    products: list[tuple],
    provider_payment_charge_id,
    paid_amount,
    fio,
    phone,
    address,
):
    if os.path.exists(ORDERS_FILE):
        wb = load_workbook(ORDERS_FILE)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(
            ['Products', 'Payment Charge ID', 'Paid Amount', 'FIO', 'Phone', 'Address', 'Paid', 'Client Username']
        )

    ws = wb.active
    products_str = '; '.join([f"ID: {pid}, Name: {info['name']}, Price: {info['price']}, Qty: {info['quantity']}"
                             for pid, info in products])

    ws.append([products_str, provider_payment_charge_id, paid_amount, fio, phone, address, 'Оплачено', client_username])
    wb.save(ORDERS_FILE)


def are_keyboards_equal(current_keyboard, new_keyboard):
    if len(current_keyboard) != len(new_keyboard):
        return False

    for row_current, row_new in zip(current_keyboard, new_keyboard):
        if len(row_current) != len(row_new):
            return False
        for btn_current, btn_new in zip(row_current, row_new):
            if (btn_current.text != btn_new.text or btn_current.callback_data != btn_new.callback_data):
                return False
    return True


def escape_markdown_v2(text: str) -> str:
    """Экранирование зарезервированных символов для MarkdownV2."""
    reserved_chars = r'([-\.!#])'
    return re.sub(reserved_chars, r'\\\1', text)
